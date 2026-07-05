"""Fill the Skyscale Media Conde Nast print plan template from publisher rate cards.

Implements the condenast-plan-filler skill (v3.0):
- text is copied verbatim from the rate cards, never paraphrased
- every rate card placement row gets its own template row — sections grow
  (rows inserted before the section total) or shrink (surplus rows removed)
  so each CNT section fits its rate card exactly
- the template's merge structure is rebuilt around the data: the rate card's
  own vertical merges (e.g. the US package/cost block) are replicated through
  the column map, and the Market / Media columns are merged per group
- section totals and the grand total are written as live SUM formulas at
  their shifted positions
"""
from __future__ import annotations

import io
import re
from copy import copy as _copy_style

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


class CNPlanError(ValueError):
    """User-facing extraction/validation error (wrong file, missing headers…)."""


# ── Template geometry (rows in the EMPTY template) ──────────────────────────
# start = first data row of the CNT section, slots = empty rows available,
# total = the section's TOTAL row (merged A:I band).
_SECTIONS = {
    'india': {'start': 10, 'slots': 3,  'total': 13, 'market': 'INDIA', 'label': 'CNT INDIA TOTAL'},
    'uk':    {'start': 14, 'slots': 6,  'total': 20, 'market': 'UK',    'label': 'CNT UK TOTAL'},
    'us':    {'start': 21, 'slots': 10, 'total': 31, 'market': 'US',    'label': 'CNT US TOTAL'},
}
# Every total row that feeds the grand total, in template coordinates:
# NETWORK18 INDIA, CNT INDIA, CNT UK, CNT US, Panrotas, EFE, ANSA.
_ALL_TOTAL_ROWS = [9, 13, 20, 31, 36, 38, 40]
_EFE_TOTAL, _EFE_DATA = 38, 37   # template has J38 = J37
_GRAND_TOTAL = 41

# Template columns: A Market, B Media, C Elements, D Format, E Platform,
# F Unit Type, G KPI's, H Buy Type, I Net CPM, J Net Total.
_MIN_COL_WIDTHS = {1: 18, 2: 22, 3: 48, 4: 35, 5: 28, 6: 22, 7: 18, 8: 12, 9: 12, 10: 14}

_ROW_KEYS = ('media', 'elements', 'format', 'platform', 'unit_type',
             'kpis', 'buy_type', 'net_cpm', 'net_total')
_WRITE_COLS = {'elements': 3, 'format': 4, 'platform': 5, 'unit_type': 6,
               'kpis': 7, 'buy_type': 8, 'net_cpm': 9, 'net_total': 10}


# ── Small helpers ────────────────────────────────────────────────────────────

def _s(v):
    """Cell value as stripped text — verbatim, '' for None."""
    return '' if v is None else str(v).strip()


def _blank(v):
    """True when a cell is effectively empty for row-detection logic."""
    return _s(v).lower() in ('', 'na', 'n/a', 'nan', 'none')


def _num(v):
    """Strict numeric value or None (used for money columns only)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = _s(v).replace(',', '')
    if re.fullmatch(r'-?\d+(\.\d+)?', s):
        return float(s)
    return None


def _numish(v):
    """Number when the cell is purely numeric, otherwise the verbatim text."""
    n = _num(v)
    if n is None:
        return _s(v)
    return int(n) if n == int(n) else round(n, 2)


def _money(v):
    n = _num(v)
    if n is None:
        return None
    return int(n) if n == int(n) else round(n, 2)


def _grab(row, col_map, key):
    ci = col_map.get(key)
    return row[ci] if ci is not None and ci < len(row) else None


def _map_headers(header_row, patterns):
    """Map canonical keys to 0-based column indices via ordered regex patterns."""
    cells = [_s(v).lower() for v in header_row]
    col_map, used = {}, set()
    for key, pat in patterns:
        for ci, h in enumerate(cells):
            if ci not in used and h and re.search(pat, h):
                col_map[key] = ci
                used.add(ci)
                break
    return col_map


def _find_header_row(ws, must_have, market):
    """Locate the rate card header row: first row containing all `must_have` names."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        cells = [_s(v).lower() for v in row]
        if all(any(h == m or h.startswith(m) for h in cells if h) for m in must_have):
            return i, list(row)
    raise CNPlanError(
        f"Could not find the header row ({', '.join(must_have)}) in the {market} rate card — "
        f"is the right file in the {market} slot?")


def _vertical_merges(ws, col_map, tcol_by_key, row_to_seq):
    """Replicate the rate card's vertical merges onto template columns.

    Returns (template_col, first_seq, last_seq) for each card merge that spans
    2+ extracted placement rows in a mapped column.
    """
    inv = {ci: key for key, ci in col_map.items()}
    out = []
    for mr in ws.merged_cells.ranges:
        if mr.min_col != mr.max_col or mr.max_row == mr.min_row:
            continue  # only single-column vertical merges carry grouping meaning
        key = inv.get(mr.min_col - 1)
        tcol = tcol_by_key.get(key)
        if tcol is None:
            continue
        seqs = [row_to_seq[r] for r in range(mr.min_row, mr.max_row + 1) if r in row_to_seq]
        if len(seqs) >= 2:
            out.append((tcol, min(seqs), max(seqs)))
    return out


# ── Market extractors ────────────────────────────────────────────────────────
# Each returns {'rows': [row-dict…], 'vmerges': [(tcol, seq_lo, seq_hi)…],
#               'header': {client/campaign/agency/contact found in the card}}

def _extract_india(ws):
    hdr_i, hdr = _find_header_row(ws, ('media', 'elements'), 'INDIA')
    col = _map_headers(hdr, [
        ('media',      r'^media'),
        ('elements',   r'^elements'),
        ('platform',   r'^platform'),
        ('format',     r'^format'),
        ('followers',  r'^followers'),
        ('net_rate',   r'^net rate'),
        ('total_net',  r'^total net'),
        ('page_views', r'^page views'),
        ('readership', r'readership'),
        ('reach',      r'\breach\b'),
    ])
    rows, row_to_seq = [], {}
    for ri, row in enumerate(ws.iter_rows(min_row=hdr_i + 1, values_only=True), hdr_i + 1):
        if _s(row[0]).lower().startswith('total net'):
            break
        elements = _s(_grab(row, col, 'elements'))
        if not elements:
            continue  # section bands (NATIVE ARTICLE / SOCIAL MEDIA / PRINT) and subtotals

        # Unit Type + KPI derived from whichever KPI column carries a value
        pv, reach = _grab(row, col, 'page_views'), _grab(row, col, 'reach')
        readership, followers = _grab(row, col, 'readership'), _grab(row, col, 'followers')
        if not _blank(pv):
            unit, kpi = 'Page Views', _numish(pv)
        elif not _blank(reach):
            unit, kpi = 'Reach', _numish(reach)
        elif not _blank(readership):
            unit, kpi = 'Quarterly Readership', f'{_s(readership)} per issue'
        elif not _blank(followers):
            unit, kpi = 'Followers', _numish(followers)
        else:
            unit, kpi = '', ''

        rate = _grab(row, col, 'net_rate')
        if _num(rate) is not None:
            buy, cpm, total = 'FIXED', 'FLAT', _money(_grab(row, col, 'total_net'))
        else:  # e.g. "Value Ad(d)" rows — still deliverables, still get a row
            txt = _s(rate)
            buy = 'Value Add' if re.search(r'value\s*ad', txt, re.I) else (txt or 'Value Add')
            cpm, total = '-', None

        row_to_seq[ri] = len(rows)
        rows.append({
            'media':    _s(_grab(row, col, 'media')),
            'elements': elements,
            'format':   _s(_grab(row, col, 'format')),
            'platform': _s(_grab(row, col, 'platform')),
            'unit_type': unit, 'kpis': kpi,
            'buy_type': buy, 'net_cpm': cpm, 'net_total': total,
        })
    if rows and not rows[0]['media']:
        rows[0]['media'] = 'CNT'
    tcols = {'elements': 3, 'format': 4, 'platform': 5, 'net_rate': 9, 'total_net': 10}
    return {'rows': rows, 'vmerges': _vertical_merges(ws, col, tcols, row_to_seq), 'header': {}}


_UK_PLATFORM_BY_SECTION = {'media': 'cntraveller.co.uk',
                           'social': 'CN Traveller Social',
                           'production': 'N/A'}


def _extract_uk(ws):
    hdr_i, hdr = _find_header_row(ws, ('placement', 'kpi guarantee'), 'UK')
    col = _map_headers(hdr, [
        ('placement',     r'^placement'),
        ('format',        r'^format'),
        ('kpi_guarantee', r'^kpi guarantee'),
        ('kpi',           r'^kpi$'),
        ('revenue',       r'^revenue'),
        ('cpm',           r'^cpm'),
        ('total_usd',     r'^total usd'),
    ])
    rows, row_to_seq, section = [], {}, 'media'
    for ri, row in enumerate(ws.iter_rows(min_row=hdr_i + 1, values_only=True), hdr_i + 1):
        joined = ' '.join(_s(v) for v in row).lower()
        if 'total investment' in joined:
            break
        placement = _s(_grab(row, col, 'placement'))
        fmt, revenue = _grab(row, col, 'format'), _grab(row, col, 'revenue')
        if placement and _blank(fmt) and _blank(revenue):
            # section band — sets platform context for the rows below it
            low = placement.lower()
            for token, name in (('social', 'social'), ('production', 'production'), ('media', 'media')):
                if token in low:
                    section = name
                    break
            continue
        if _blank(fmt) and _blank(revenue):
            continue
        row_to_seq[ri] = len(rows)
        rows.append({
            'media':    'CNT' if not rows else '',
            'elements': placement,
            'format':   _s(fmt),
            'platform': _UK_PLATFORM_BY_SECTION[section],
            'unit_type': _s(_grab(row, col, 'kpi')),
            'kpis':     _numish(_grab(row, col, 'kpi_guarantee')),
            'buy_type': _s(revenue),
            'net_cpm':  _numish(_grab(row, col, 'cpm')),
            'net_total': _money(_grab(row, col, 'total_usd')),
        })
    # Header block (CLIENT / CAMPAIGN NAME / AGENCY / CONTACT above the table)
    header = {}
    for row in ws.iter_rows(min_row=1, max_row=hdr_i - 1, values_only=True):
        label, value = _s(row[0]).lower().rstrip(':'), _s(row[1] if len(row) > 1 else None)
        if not value:
            continue
        if label == 'client':
            header['client'] = value
        elif label == 'campaign name':
            header['campaign'] = value
        elif label == 'agency' and not _blank(value):
            header['agency'] = value
        elif label == 'contact':
            header['contact'] = value
    tcols = {'placement': 3, 'format': 4, 'kpi': 6, 'kpi_guarantee': 7,
             'revenue': 8, 'cpm': 9, 'total_usd': 10}
    return {'rows': rows, 'vmerges': _vertical_merges(ws, col, tcols, row_to_seq), 'header': header}


def _extract_us(ws):
    hdr_i, hdr = _find_header_row(ws, ('site', 'placement name'), 'US')
    col = _map_headers(hdr, [
        ('site',        r'^site$'),
        ('package',     r'^package'),
        ('placement',   r'^placement name'),
        ('platform',    r'^platform'),
        ('unit_type',   r'^unit type'),
        ('units',       r'^units$'),
        ('cost_method', r'^cost method'),
        ('rate',        r'^rate$'),
        ('cost',        r'^cost$'),
    ])
    rows, row_to_seq = [], {}
    for ri, row in enumerate(ws.iter_rows(min_row=hdr_i + 1, values_only=True), hdr_i + 1):
        site_raw = _s(_grab(row, col, 'site'))
        if 'grand total' in site_raw.lower():
            break
        placement = _s(_grab(row, col, 'placement'))
        if not placement:
            continue
        site = re.sub(r'\s+', ' ', site_raw)
        if not rows:
            media = 'CNT'  # first block of the CNT plan
        elif 'native tile driver' in placement.lower():
            media = 'Conde Nast Network'  # network-wide roadblock row
        else:
            media = site
        row_to_seq[ri] = len(rows)
        rows.append({
            'media':    media,
            'elements': _s(_grab(row, col, 'package')),   # full package text, verbatim
            'format':   placement,
            'platform': _s(_grab(row, col, 'platform')),
            'unit_type': _s(_grab(row, col, 'unit_type')),
            'kpis':     _numish(_grab(row, col, 'units')),
            'buy_type': _s(_grab(row, col, 'cost_method')),
            'net_cpm':  _numish(_grab(row, col, 'rate')),
            'net_total': _money(_grab(row, col, 'cost')),
        })
    # PRESENTED TO / CAMPAIGN block above the table
    header = {}
    for row in ws.iter_rows(min_row=1, max_row=hdr_i - 1, values_only=True):
        for ci, v in enumerate(row[:6]):
            label = _s(v).lower().rstrip(':')
            value = _s(row[ci + 1] if ci + 1 < len(row) else None)
            if label == 'presented to' and value:
                header['client'] = value.title() if value.isupper() else value
            elif label == 'campaign' and value:
                header['campaign'] = value
    tcols = {'package': 3, 'placement': 4, 'platform': 5, 'unit_type': 6,
             'units': 7, 'cost_method': 8, 'rate': 9, 'cost': 10}
    return {'rows': rows, 'vmerges': _vertical_merges(ws, col, tcols, row_to_seq), 'header': header}


_EXTRACTORS = {'india': _extract_india, 'uk': _extract_uk, 'us': _extract_us}


# ── Merge bookkeeping across row inserts/deletes ─────────────────────────────

def _shift_merges_insert(merges, at, k):
    return [(r1 + k if r1 >= at else r1, c1, r2 + k if r2 >= at else r2, c2)
            for (r1, c1, r2, c2) in merges]


def _shift_merges_delete(merges, ds, k):
    de = ds + k - 1
    out = []
    for (r1, c1, r2, c2) in merges:
        overlap = max(0, min(r2, de) - max(r1, ds) + 1)
        remaining = (r2 - r1 + 1) - overlap
        if remaining <= 0:
            continue
        nr1 = r1 - k if r1 > de else (ds if r1 >= ds else r1)
        out.append((nr1, c1, nr1 + remaining - 1, c2))
    return out


# ── Main fill ────────────────────────────────────────────────────────────────

def fill_cn_print_plan(template_source, cards, header=None):
    """Fill the CN print plan template.

    template_source: path or file-like of the EMPTY template.
    cards: {'india'|'uk'|'us': xlsx bytes} — any subset.
    header: {'client','campaign','agency','contact'} overrides (optional).
    Returns a BytesIO of the filled workbook.
    """
    header = {k: v for k, v in (header or {}).items() if v}

    extracted = {}
    for mkt, blob in cards.items():
        card_wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
        data = _EXTRACTORS[mkt](card_wb.active)
        if not data['rows']:
            raise CNPlanError(
                f'No placement rows found in the {_SECTIONS[mkt]["market"]} rate card — '
                'is the right file in that slot?')
        extracted[mkt] = data
        for k, v in data['header'].items():
            header.setdefault(k, v)

    wb = openpyxl.load_workbook(template_source)
    ws = wb.active

    merges = [(mr.min_row, mr.min_col, mr.max_row, mr.max_col)
              for mr in ws.merged_cells.ranges]
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    # Resize each filled section to its rate card, bottom-up so row numbers of
    # the sections still to process are unaffected. openpyxl does not shift
    # merged ranges on insert/delete, so we track them manually.
    deltas = {m: 0 for m in _SECTIONS}
    for mkt in ('us', 'uk', 'india'):
        if mkt not in extracted:
            continue
        sec = _SECTIONS[mkt]
        n = len(extracted[mkt]['rows'])
        delta = n - sec['slots']
        if delta > 0:
            ws.insert_rows(sec['total'], delta)
            merges = _shift_merges_insert(merges, sec['total'], delta)
            # inherit style + height from the section's last original slot row
            src = sec['total'] - 1
            for r in range(sec['total'], sec['total'] + delta):
                ws.row_dimensions[r].height = ws.row_dimensions[src].height
                for c in range(1, 11):
                    ws.cell(r, c)._style = _copy_style(ws.cell(src, c)._style)
        elif delta < 0:
            ws.delete_rows(sec['start'] + n, -delta)
            merges = _shift_merges_delete(merges, sec['start'] + n, -delta)
        deltas[mkt] = delta

    def new_pos(template_row):
        return template_row + sum(d for m, d in deltas.items()
                                  if _SECTIONS[m]['total'] <= template_row)

    # Data areas of the filled sections (final coordinates) — the template's
    # own merges inside these are dropped and rebuilt from the rate cards.
    areas = {}
    for mkt in extracted:
        sec = _SECTIONS[mkt]
        start = new_pos(sec['total']) - len(extracted[mkt]['rows'])
        areas[mkt] = (start, new_pos(sec['total']) - 1)

    def in_filled_area(r1, r2):
        return any(a <= r1 and r2 <= b for a, b in areas.values())

    # ── Write values ──
    for key, row_i in (('client', 1), ('campaign', 2), ('agency', 3), ('contact', 4)):
        ws.cell(row_i, 2, header.get(key, ''))

    wrap_cells = []
    for mkt, data in extracted.items():
        sec = _SECTIONS[mkt]
        start, end = areas[mkt]
        ws.cell(start, 1, sec['market'])
        prev_media = None
        for i, r in enumerate(data['rows']):
            rr = start + i
            if r['media'] and r['media'] != prev_media:
                ws.cell(rr, 2, r['media'])
                prev_media = r['media']
            for key, colx in _WRITE_COLS.items():
                v = r[key]
                if v is None or v == '':
                    continue
                cell = ws.cell(rr, colx, v)
                if isinstance(v, str) and (len(v) > 30 or '\n' in v):
                    wrap_cells.append(cell)
        trow = new_pos(sec['total'])
        ws.cell(trow, 1, sec['label'])
        ws.cell(trow, 10, f'=SUM(J{start}:J{end})')

    # Rebuild the formulas the row shifts invalidated
    ws.cell(new_pos(_EFE_TOTAL), 10, f'=J{new_pos(_EFE_DATA)}')
    refs = ','.join(f'J{new_pos(r)}' for r in _ALL_TOTAL_ROWS)
    ws.cell(new_pos(_GRAND_TOTAL), 10, f'=SUM({refs})')

    # ── Re-merge ──
    # 1. surviving template merges (shifted), except inside rebuilt data areas
    for (r1, c1, r2, c2) in merges:
        if in_filled_area(r1, r2):
            continue
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    # 2. Market column across each filled section
    for mkt, (start, end) in areas.items():
        if end > start:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        # 3. Media column per group of consecutive rows (forward-filled, so a
        #    repeated media name stays in the same group as the rows above it)
        ff, cur = [], ''
        for r in extracted[mkt]['rows']:
            cur = r['media'] or cur
            ff.append(cur)
        g0 = 0
        for i in range(1, len(ff) + 1):
            if i == len(ff) or ff[i] != ff[g0]:
                if i - 1 > g0 and ff[g0]:
                    ws.merge_cells(start_row=start + g0, start_column=2,
                                   end_row=start + i - 1, end_column=2)
                g0 = i
        # 4. the rate card's own vertical merges, mapped through the column map
        for (tcol, lo, hi) in extracted[mkt]['vmerges']:
            ws.merge_cells(start_row=start + lo, start_column=tcol,
                           end_row=start + hi, end_column=tcol)

    # ── Presentation: wrap long text, keep columns wide enough, size rows ──
    for cell in wrap_cells:
        base = cell.alignment or Alignment()
        cell.alignment = Alignment(horizontal=base.horizontal, vertical='top', wrap_text=True)
    for colx, width in _MIN_COL_WIDTHS.items():
        letter = get_column_letter(colx)
        cur = ws.column_dimensions[letter].width
        if cur is None or cur < width:
            ws.column_dimensions[letter].width = width
    for mkt, (start, end) in areas.items():
        for rr in range(start, end + 1):
            longest = max((len(line)
                           for c in ws[rr][:10] if isinstance(c.value, str)
                           for line in c.value.split('\n')), default=0)
            n_lines = max((str(c.value).count('\n') + 1
                           for c in ws[rr][:10] if isinstance(c.value, str)), default=1)
            est = max(n_lines, -(-longest // 45))  # ceil(longest / wrapped width)
            if est > 1:
                ws.row_dimensions[rr].height = min(150, 15 * est)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
